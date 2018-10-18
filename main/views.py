"""
Views here are mostly subclasses of the Django generic CreateView. Each view has
its own template, named in snake-case for the model to-be-created. Each view
presents a Django form based on a model.

The views are linked together in a linear sequence that reflects the
dependencies of the data. Successfully submitting one form redirects the user to
the next form. Each model has a foreign key into the preceding model. One
exception is PrimerDesign which depends on GuideSelection, two steps back in the
sequence.
"""
import logging
import os
import time

from concurrent.futures import ThreadPoolExecutor
from io import BytesIO, StringIO
from itertools import islice
from typing import Any

import openpyxl  # type: ignore
import sample_sheet as illumina  # type: ignore

from django.http import Http404
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.views import View
from django.views.generic import DetailView, ListView
from django.views.generic.edit import CreateView

import webscraperequest

from crispresso.fastqs import find_matching_pairs
from crispresso.s3 import download_fastqs
from protospacex import get_cds_chr_loc, get_cds_seq

from main import samplesheet
from main.forms import *
from main.models import *
from utils import conversions
from utils.validators import is_ensemble_transcript

# TODO (gdingle): move somewhere better
CRISPRESSO_ROOT_URL = 'http://crispresso:5000/'
CRISPRESSO_PUBLIC_ROOT_URL = 'http://0.0.0.0:5000/'

logger = logging.getLogger(__name__)


class IndexView(ListView):
    model = Experiment
    template_name = 'index.html'

    # TODO (gdingle): fill in
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['analyses'] = (a for a in Analysis.objects.all() if a.is_complete)
        return context


def index(request):
    # TODO (gdingle): create useful index
    return HttpResponse("Hello, world. You're at the main index.")


# TODO (gdingle): make CreateUpdateView
# see https://stackoverflow.com/a/48116803
class CreatePlusView(CreateView):
    """
    Simplifies adding foreign keys and other pre-determined data to a ModelForm
    before saving it.

    See https://github.com/django/django/blob/master/django/views/generic/edit.py.
    """

    # TODO (gdingle): handle exceptions here to form
    # see form_invalid, non_field_errors, https://docs.djangoproject.com/en/2.1/topics/class-based-views/generic-editing/
    def form_valid(self, form: ModelForm) -> HttpResponse:
        obj = form.save(commit=False)
        obj = self.plus(obj)
        obj.save()
        self.object = obj
        return HttpResponseRedirect(self.get_success_url())

    def plus(self, obj: models.Model) -> models.Model:
        """Add attributes to a model before saving it."""


class ExperimentView(CreateView):
    template_name = 'experiment.html'
    form_class = ExperimentForm
    success_url = '/main/experiment/{id}/guide-design/'


class GuideDesignView(CreatePlusView):
    template_name = 'guide-design.html'
    form_class = GuideDesignForm
    success_url = '/main/guide-design/{id}/progress/'

    def _normalize_targets(self, targets, guide_design):
        genome = guide_design.genome
        if all(is_seq(t) for t in targets):
            # TODO (gdingle): gggenome all seqs
            return targets

        if guide_design.hdr_tag:
            assert all(is_ensemble_transcript(t) for t in targets), 'must be ENST for HDR'
            assert genome == 'hg38', 'only implemented for hg38'
            if guide_design.hdr_tag == 'per_target':
                func = get_cds_chr_loc
                cds_indexes = guide_design.cds_index
                cds_lengths = guide_design.cds_length
                with ThreadPoolExecutor() as pool:
                    return list(pool.map(func, targets, cds_indexes, cds_lengths))
            else:
                func = functools.partial(
                    get_cds_chr_loc,
                    cds_index=guide_design.cds_index,
                    length=guide_design.cds_length)
                with ThreadPoolExecutor() as pool:
                    return list(pool.map(func, targets))
        elif all(is_gene(t) or is_ensemble_transcript(t) for t in targets):
            # TODO (gdingle): this still needs some work to get best region of gene and not the whole thing
            func = functools.partial(
                conversions.gene_to_chr_loc,
                genome=genome)
            with ThreadPoolExecutor() as pool:
                return list(pool.map(func, targets))

        assert False

    def _get_target_seqs(self, targets, guide_design):
        genome = guide_design.genome
        if all(is_seq(t) for t in targets):
            return targets

        if guide_design.hdr_tag:
            assert genome == 'hg38', 'only implemented for hg38'
            if guide_design.hdr_tag == 'per_target':
                func = get_cds_seq
                cds_indexes = guide_design.cds_index
                cds_lengths = guide_design.cds_length
                with ThreadPoolExecutor() as pool:
                    return list(pool.map(func, targets, cds_indexes, cds_lengths))
            else:
                func = functools.partial(
                    get_cds_seq,
                    cds_index=guide_design.cds_index,
                    length=guide_design.cds_length)
                with ThreadPoolExecutor() as pool:
                    return list(pool.map(func, targets))
        elif all(is_gene(t) or is_ensemble_transcript(t) for t in targets):
            # TODO (gdingle): get gene seq or disable in UI
            assert False, 'not implemented'
        else:
            func = functools.partial(
                conversions.chr_loc_to_seq,
                genome=genome)
            with ThreadPoolExecutor() as pool:
                return list(pool.map(func, targets))

        assert False

    def plus(self, obj):
        obj.experiment = Experiment.objects.get(id=self.kwargs['id'])

        targets_cleaned, target_tags = obj.parse_targets_raw()
        obj.target_tags = target_tags

        obj.targets = self._normalize_targets(
            targets_cleaned,
            obj
        )
        obj.target_seqs = self._get_target_seqs(
            targets_cleaned,
            obj
        )

        batch = webscraperequest.CrisporGuideBatchWebRequest(obj)
        pre_filter = obj.wells_per_target * 5  # 5 based on safe-harbor experiment
        largs = [[target_seq, obj.experiment.name, obj.genome, obj.pam, target, pre_filter]
                 for target, target_seq in zip(obj.targets, obj.target_seqs)]
        batch.start(largs, [-2])

        return obj


class GuideDesignProgressView(View):

    template_name = 'guide-design-progress.html'
    success_url = '/main/guide-design/{id}/guide-selection/'

    def get(self, request, **kwargs):
        guide_design = GuideDesign.objects.get(id=self.kwargs['id'])
        batch_status = webscraperequest.CrisporGuideBatchWebRequest(guide_design).get_batch_status()

        if not batch_status.is_successful:
            return render(request, self.template_name, locals())
        else:
            return HttpResponseRedirect(
                self.success_url.format(id=self.kwargs['id']))


class GuideSelectionView(CreatePlusView):
    template_name = 'guide-selection.html'
    form_class = GuideSelectionForm
    success_url = '/main/guide-selection/{id}/primer-design/'

    @staticmethod
    def _top_guides(guide_data, guide_design, min_score=30):
        """
        Select the top guides for further manual selection depending on a few
        vars, such as HDR and wells_per_target.
        """
        by = 'distance' if guide_design.hdr_tag else 'score'
        top = guide_design.wells_per_target

        guide_seqs = guide_data['guide_seqs']
        if not guide_seqs or guide_seqs.get('not found'):
            return guide_seqs

        # First score should be the MIT specificity score
        # Filter out 'Not found'
        scores = dict((k, int(s[0]))
                      for k, s in guide_data['scores'].items()
                      if s[0].isdigit())

        if by == 'distance':
            guide_seqs = dict(g for g in guide_seqs.items()
                              # TODO (gdingle): is 30 (yellow) a good cut-off?
                              if scores.get(g[0], 0) > min_score)

        def func(t):
            if by == 'score':
                return int(scores[t[0]])
            elif by == 'distance':
                # s29+, s26+, etc
                # change dist depending on codon target
                # TODO (gdingle): what about forward and reverse?
                # we should really rank by cut to insert distance
                guide_offset = int(t[0][1:-1])
                if guide_design.cds_index == -1:
                    # Reproduce protospacex min_length fetch behavior
                    mid = guide_design.cds_length // 2
                    return abs(mid - guide_offset)
                else:
                    return guide_offset

        guide_seqs = sorted(
            guide_seqs.items(),
            key=func,
            reverse=(by == 'score'))
        return dict(islice(guide_seqs, top))

    def get_initial(self):
        guide_design = GuideDesign.objects.get(id=self.kwargs['id'])
        return {
            'selected_guides': dict(
                (g['target'], self._top_guides(g, guide_design))
                for g in guide_design.guide_data),
        }

    def plus(self, obj):
        obj.guide_design = GuideDesign.objects.get(id=self.kwargs['id'])
        return obj

    def get_context_data(self, **kwargs):
        guide_design = GuideDesign.objects.get(id=self.kwargs['id'])
        kwargs['crispor_urls'] = dict(
            (gd['target'], gd['url'])
            for gd in guide_design.guide_data
            if gd.get('url'))
        return super().get_context_data(**kwargs)


class PrimerDesignView(CreatePlusView):
    template_name = 'primer-design.html'
    form_class = PrimerDesignForm
    success_url = '/main/primer-design/{id}/progress/'

    def plus(self, obj):
        guide_selection = GuideSelection.objects.get(id=self.kwargs['id'])
        obj.guide_selection = guide_selection

        sheet = samplesheet.from_guide_selection(guide_selection)
        batch = webscraperequest.CrisporPrimerBatchWebRequest(obj)
        largs = [[row['_crispor_batch_id'],
                  row['_crispor_pam_id'],
                  obj.amplicon_length,
                  obj.primer_temp,
                  guide_selection.guide_design.pam,
                  row['_crispor_guide_id']]
                 for row in sheet.to_records()]
        batch.start(largs, [-1])

        # TODO (gdingle): run crispr-primer if HDR experiment
        # https://github.com/chanzuckerberg/crispr-primer
        return obj


class PrimerDesignProgressView(View):
    template_name = 'primer-design-progress.html'
    success_url = '/main/primer-design/{id}/primer-selection/'

    def get(self, request, **kwargs):
        primer_design = PrimerDesign.objects.get(id=kwargs['id'])
        batch_status = webscraperequest.CrisporPrimerBatchWebRequest(
            primer_design).get_batch_status()

        if not batch_status.is_successful:
            return render(request, self.template_name, locals())
        else:
            return HttpResponseRedirect(
                self.success_url.format(id=self.kwargs['id']))


class PrimerSelectionView(CreatePlusView):
    template_name = 'primer-selection.html'
    form_class = PrimerSelectionForm
    success_url = '/main/primer-selection/{id}/experiment-summary/'

    def get_initial(self):
        primer_data = PrimerDesign.objects.get(id=self.kwargs['id']).primer_data

        def get_fwd_and_rev_primers(ontarget_primers: dict):
            values = list(ontarget_primers.values())
            if not values:
                return 'not found'
            # TODO (gdingle): _transform_primer_product here instead of later?
            return values[0], values[1]

        return {
            'selected_primers': dict(
                (p['target'], get_fwd_and_rev_primers(p['ontarget_primers'])
                 if p['success'] else p['error'])
                for p in primer_data)
        }

    def plus(self, obj):
        obj.primer_design = PrimerDesign.objects.get(id=self.kwargs['id'])
        return obj

    def get_context_data(self, **kwargs):
        primer_data = PrimerDesign.objects.get(id=self.kwargs['id']).primer_data
        kwargs['crispor_urls'] = dict(
            (p['target'], p['url'] + '#ontargetPcr')
            for p in primer_data)
        return super().get_context_data(**kwargs)


class ExperimentSummaryView(View):
    template_name = 'experiment-summary.html'

    def get(self, request, *args, **kwargs):
        try:
            primer_selection = PrimerSelection.objects.get(id=kwargs['id'])
        except PrimerSelection.DoesNotExist:
            # Alternate path to summary
            try:
                primer_selection = PrimerSelection.objects.filter(
                    primer_design__guide_selection__guide_design__experiment=kwargs['id'])[0]
            except IndexError:
                raise Http404('Experiment summary does not exist')

        sheet = samplesheet.from_primer_selection(primer_selection)
        sheet = self._prepare_sheet(sheet)

        primer_design = primer_selection.primer_design
        guide_selection = primer_design.guide_selection
        guide_design = guide_selection.guide_design
        experiment = guide_design.experiment

        if request.GET.get('download') == 'xls':
            del sheet['Well Pos']  # already contained in index
            excel_file = samplesheet.to_excel(sheet)
            title = experiment.name + ' summary'
            return _excel_download_response(excel_file, title)

        # TODO (gdingle): download csv

        # max length to show of table cell values
        # 26 is optimized for laptop screen and chr loc
        show = request.GET.get('show', 26)

        return render(request, self.template_name, locals())

    def _prepare_sheet(self, sheet):
        """Modify sheet for optimal rendering"""
        sheet = sheet.loc[:, 'target_input':]  # type: ignore
        sheet = sheet.loc[:, [not c.startswith('_') for c in sheet.columns]]
        sheet = sheet.dropna(axis=1, how='all')
        sheet.insert(0, 'well_pos', sheet.index)
        sheet.insert(1, 'well_num', range(1, len(sheet) + 1))
        sheet.columns = [c.replace('_', ' ').title() for c in sheet.columns]
        return sheet


class AnalysisView(CreatePlusView):
    template_name = 'analysis.html'
    form_class = AnalysisForm
    success_url = '/main/analysis/{id}/progress/'

    def plus(self, obj):
        # TODO (gdingle): use predetermined s3 location of fastq
        fastqs = download_fastqs(obj.s3_bucket, obj.s3_prefix, overwrite=False)
        assert len(fastqs) <= 384, 'Fastqs should be from max one plate'

        # Redirect to intermediate page if custom analysis
        if obj.is_custom:
            self.success_url = '/main/analysis/{id}/custom/'
            obj.fastq_data = fastqs
            return obj

        sheet = samplesheet.from_analysis(obj)

        # TODO (gdingle): create dir per download, as in seqbot
        obj.fastq_data = find_matching_pairs(fastqs, sheet.to_records())

        sheet = samplesheet.from_analysis(obj)

        webscraperequest.CrispressoBatchWebRequest.start_analysis(
            obj, sheet.to_records())
        return obj


class CustomAnalysisView(View):

    template_name = 'custom-analysis.html'
    success_url = '/main/analysis/{id}/progress/'
    form = CustomAnalysisForm

    def get(self, request, **kwargs):
        analysis = Analysis.objects.get(id=kwargs['id'])
        return render(request, self.template_name, {
            **kwargs,
            'form': self.form(),
            'analysis': analysis,
        })

    def post(self, request, **kwargs):
        analysis = Analysis.objects.get(id=kwargs['id'])
        form = self.form(request.POST, request.FILES)

        if not form.is_valid():
            return render(request, self.template_name, {
                **kwargs,
                'form': self.form(),
                'analysis': analysis,
            })

        file = form.cleaned_data['file']
        sheet = samplesheet.from_excel(file)
        fastq_data = analysis.fastq_data
        assert len(fastq_data), 'Fastqs must be present for a custom analysis'

        # fastq_data is initially a flat list. Process only on initial post
        # because find_matching_pairs is expensive.
        if isinstance(fastq_data[0], str):
            fastq_data = find_matching_pairs(fastq_data, sheet.to_records())

        sheet['fastq_fwd'] = [pair[0] for pair in fastq_data]
        sheet['fastq_rev'] = [pair[1] for pair in fastq_data]

        # Save only after sheet is validated above
        analysis.fastq_data = fastq_data
        analysis.save()

        webscraperequest.CrispressoBatchWebRequest.start_analysis(
            analysis, sheet.to_records())

        return HttpResponseRedirect(
            self.success_url.format(id=self.kwargs['id']))


class AnalysisProgressView(View):
    template_name = 'analysis-progress.html'
    success_url = '/main/analysis/{id}/results/'

    def get(self, request, **kwargs):
        analysis = Analysis.objects.get(id=kwargs['id'])
        batch_status = webscraperequest.CrispressoBatchWebRequest(analysis).get_batch_status()

        if not batch_status.is_successful:
            return render(request, self.template_name, locals())
        else:
            return HttpResponseRedirect(
                self.success_url.format(id=self.kwargs['id']))


class ResultsView(View):
    template_name = 'crispresso-results.html'

    def get(self, request, *args, **kwargs):
        analysis = Analysis.objects.get(id=self.kwargs['id'])
        try:
            if analysis.is_custom:
                sheet = samplesheet.from_custom_analysis(analysis)
                input_data = [
                    (d['input_data'],
                        [f.split('/')[-1] for f in d['input_files']])
                    for d in analysis.results_data]
            else:
                sheet = samplesheet.from_analysis(analysis)
            # TODO (gdingle): mode to show each stat col as % of total?
        except IndexError:
            raise Http404('Analysis results do not exist')
        return render(request, self.template_name, locals())


# TODO (gdingle): is DetailView needed here?
class OrderFormView(DetailView):
    """
    Produces a downloadable Excel order form for IDT. The model must have a
    plate layout.
    """

    model: models.Model = None
    seq_key: str

    def _create_excel_file(self, sheet: samplesheet.pandas.DataFrame, title: str):
        # TODO (gdingle): use pandas to_excel here instead?
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = title[0:31]  # Excel limits to 30 chars
        ws['A1'] = 'Well Position'
        ws['B1'] = 'Sequence Name'
        ws['C1'] = 'Sequence'

        for j, well_pos in enumerate(sheet.index):
            for i, seq_key in enumerate(self.seq_keys):
                index = str((j * len(self.seq_keys)) + i + 2)
                ws['A' + index] = well_pos
                # TODO (gdingle): is this a good name for each sequence?
                row = sheet.loc[well_pos]
                ws['B' + index] = '{} {}'.format(
                    row['_crispor_guide_id'], seq_key,
                )
                ws['C' + index] = row[seq_key]

        return openpyxl.writer.excel.save_virtual_workbook(wb)

    def get(self, request, *args, **kwargs):
        instance = self.model.objects.get(id=kwargs['id'])
        # TODO (gdingle): friendlier title?
        title = request.path.replace('/', ' ').replace('main ', '')
        excel_file = self._create_excel_file(instance.samplesheet, title)
        return _excel_download_response(excel_file, title)


class GuideOrderFormView(OrderFormView):

    model = GuideSelection
    seq_keys = ('guide_seq',)


class PrimerOrderFormView(OrderFormView):

    model = PrimerSelection
    seq_keys = ('primer_seq_fwd', 'primer_seq_rev')


class IlluminaSheetView(View):

    def _make_sample(self, experiment: Experiment, row: Any) -> illumina.Sample:
        return illumina.Sample({
            # A short ID assigned to the specific study and/or project based
            #  on the conventions used at your institution/group. Accepted
            #  characters include numbers, letters, "-", and "_".
            'Study_ID': experiment.short_name,

            # Description of a project to which the current sequencing run
            # belongs. Special characters and commas are not allowed.
            # (Required)
            # e.g. Understanding differences in RNA expression
            # between pancreatic cells collected from healthy pancreas and
            # patients with pancreatic cancer.
            'Study_Description': experiment.description,

            # Description (longer than 10 characters) of the specific
            # biological sample. The description should be the same for all
            # the samples derived from this BioSample. Special characters
            # and commas are not allowed.
            # (Required)
            # e.g. Microbiome sample collected on Feb. 30 2017 from mouse 002.
            # e.g. Liver tissue obtained on Feb. 30 2017 from patient 012.
            # Leave blank, to be filled in by researcher
            'BioSample_ID': None,

            # Description (longer than 10 characters) of the specific
            # biological sample. The description should be the same for all
            # the samples derived from this BioSample. Special characters
            # and commas are not allowed.
            # (Required) e.g. Microbiome sample
            # collected on Feb. 30 2017 from mouse 002. e.g. Liver tissue
            # obtained on Feb. 30 2017 from patient 012.
            # Leave blank, to be filled in by researcher
            'BioSample_Description': None,

            # A short ID assigned to the specific library based on the
            # conventions used at your institution/group. Accepted
            # characters include numbers, letters, "-", and "_". This field
            # should be unique for each row. Sample_ID can be the same as
            # Sample_Name
            # (Required)
            'Sample_ID': experiment.short_name + '-' + row.index,

            # A distinct and descriptive name for each specific library.
            # Accepted characters are numbers, letters, "-", and "_". Name
            # must begin with a letter. This field should be unique for each
            # row. The sample name will go into the final fastq file names.
            # (Required) e.g. Mouse_Liver_SingleCell_plate02_A10_20171210
            'Sample_Name': experiment.short_name + '-' + row.index,

            # If people are combining samples to sequence on the same run,
            # this column is used to keep track of to whom each sample
            # belongs to. Please fill in following the format
            # FirstName_LastName. If left blank, we will use submitter's
            # information.
            # e.g. Stephen_Quake
            'Sample_Owner': experiment.researcher.full_name.replace(' ', '_'),

            # Name of the first index. Accepted characters are numbers,
            # letters, "-", and "_"
            # (Required)
            # e.g. Nextera - N700
            # Leave blank, to be filled in by researcher
            'Index_ID': None,

            # Sequence of the first index
            # (Required)
            # e.g. ATAGCGCT
            # Leave blank, to be filled in by researcher
            'Index': None,

            # Name of the second index. Accepted characters are numbers,
            # letters, "-", and "_"
            # (Required) e.g. Nextera - S500
            # Leave blank, to be filled in by researcher
            'Index2_ID': None,

            # Sequence of the second index
            # (Required) e.g. ATAGCGCT
            # Leave blank, to be filled in by researcher
            'Index2': None,

            # This field is used to record the organism the DNA comes from.
            # Some examples are Human, Mouse, Mosquito, Yeast, Bacteria,
            # etc.
            # (Required)
            'Organism': GuideDesign.GENOME_TO_ORGANISM[row['target_genome']],

            # TODO (gdingle): will these below ever be useful?
            # This field is used record the host of the organism where the
            # DNA originated, if applicable. For example, if the Organism is
            # Fungus and the Fungus is isolated from a Human fecal sample,
            # then the Host is Human. If the organism is Human, then host is
            # left blank.
            'Host': None,

            # Gender of the organism (if applicable) e.g. M or F
            'Gender': None,

            # Tissue origin of the particular sample (if applicable) e.g. Liver
            'Tissue_Source': None,

            # FACS markers used delimited by a semicolon ";" with no space.
            # Accepted characters are numbers, letters, "-" and "_". e.g.
            # Epcam;CD45
            'FACS_Markers': None,
        })

    def get(self, request, **kwargs):
        # TODO (gdingle): do we actually need primer selection here? or is guide enough?
        primer_selection = PrimerSelection.objects.get(id=kwargs['id'])
        sheet = samplesheet.from_primer_selection(primer_selection)
        experiment = primer_selection.primer_design.guide_selection.guide_design.experiment

        illumina_sheet = illumina.SampleSheet()
        # TODO (gdingle): Add link to Biohub submission form
        illumina_sheet.Header['Crispycrunch'] = 'Please fill in the following required columns: BioSample_ID, BioSample_Description, Index_ID, Index, Index2_ID, Index2.'
        for row in sheet.to_records():
            illumina_sheet.add_sample(self._make_sample(experiment, row))

        csv = StringIO()
        illumina_sheet.write(csv)
        filename = f'Illumina sample sheet for experiment {experiment.name}'
        response = HttpResponse(csv.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="{}.csv"'.format(filename)

        return response


def _excel_download_response(excel_file: BytesIO, title: str) -> HttpResponse:
    response = HttpResponse(
        excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="{}.xlsx"'.format(title)
    return response
